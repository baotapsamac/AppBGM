# -*- coding: utf-8 -*-
"""
Lõi xử lý: đọc placeholder {{Tên trường}} / {{Tên trường:định dạng}} trong .docx,
đọc dữ liệu nhiều cột (nhiều bài giảng) trong .xlsx, và gộp lại thành .docx hoàn thiện
mà không phá vỡ định dạng gốc (chỉ sửa nội dung text bên trong các thẻ <w:t> có sẵn).
"""
import re
import copy
import zipfile
import shutil
import tempfile
import os
import subprocess
from datetime import datetime, date
from lxml import etree
from openpyxl import load_workbook

W = '{http://schemas.openxmlformats.org/wordprocessingml/2006/main}'
XMLNS = '{http://www.w3.org/XML/1998/namespace}'
PLACEHOLDER_RE = re.compile(r'\{\{([^{}:]+)(?::([^{}]*))?\}\}')

XML_PARTS = ['word/document.xml']  # có thể mở rộng thêm header/footer nếu cần


# ---------------------------------------------------------------- Đọc docx --
def _iter_xml_parts(docx_path):
    with zipfile.ZipFile(docx_path) as z:
        names = z.namelist()
        parts = [n for n in names if n == 'word/document.xml']
        parts += [n for n in names if re.match(r'word/(header|footer)\d*\.xml', n)]
        for n in parts:
            yield n, z.read(n)


def extract_fields(docx_path):
    """Trả về (fields_meta, field_formats):
    - fields_meta: [(ten_truong, dinh_dang_dai_dien)] theo thứ tự xuất hiện đầu tiên,
      dùng để quyết định kind/định dạng cho ô nhập liệu trên giao diện.
    - field_formats: {ten_truong: {dinh_dang1, dinh_dang2, ...}} — một trường có thể
      được đặt trong nhiều vị trí khác nhau của văn bản với các định dạng ngày khác
      nhau (ví dụ vừa "THÁNG m NĂM yyyy" vừa "Ngày dd tháng m năm yyyy"); tập hợp này
      giúp merge_docx() tính đúng giá trị hiển thị cho từng vị trí thay vì chỉ dùng
      chung một định dạng đại diện cho mọi chỗ.
    """
    seen = {}
    order = []
    all_formats = {}
    for _, raw in _iter_xml_parts(docx_path):
        root = etree.fromstring(raw)
        full = ''.join(t.text or '' for t in root.iter(W + 't'))
        for m in PLACEHOLDER_RE.finditer(full):
            name, fmt = m.group(1).strip(), m.group(2)
            if name not in seen:
                seen[name] = fmt
                order.append(name)
            elif fmt and not seen[name]:
                seen[name] = fmt
            if fmt:
                all_formats.setdefault(name, set()).add(fmt)
    fields_meta = [(k, seen[k]) for k in order]
    return fields_meta, all_formats


def field_kind(name, fmt):
    if fmt:
        return 'date'
    if name.startswith('Thời gian'):
        return 'number'
    return 'text'


# ---------------------------------------------------------------- Đọc xlsx --
def read_lectures(xlsx_path):
    """
    Đọc dữ liệu bài giảng từ file Excel.
    Hỗ trợ nạp cả file xuất từ AppBGM (có hàng tiêu đề nhóm và tô màu) và file Excel thông thường.
    """
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    max_col = ws.max_column
    max_row = ws.max_row

    SECTION_TITLES = [
        'THÔNG TIN HÀNH CHÍNH', 'MỤC ĐÍCH & YÊU CẦU', 'NỘI DUNG & PHÂN BỔ',
        'TỔ CHỨC, PHƯƠNG PHÁP', 'CHI TIẾT CÁC PHẦN', 'TÊN TRƯỜNG THÔNG TIN',
        'PHẦN I', 'PHẦN II', 'PHẦN III', 'PHẦN IV', 'PHẦN V',
        'NHÓM I', 'NHÓM II', 'NHÓM III', 'NHÓM IV', 'NHÓM V'
    ]

    field_order = []
    field_rows = []

    for r in range(2, max_row + 1):
        v = ws.cell(r, 1).value
        if v is None:
            continue
        val_str = str(v).strip()
        if not val_str:
            continue
        if any(st in val_str.upper() for st in SECTION_TITLES) or val_str.startswith(('🏛️', '🎯', '⏱️', '🏫', '📝')):
            continue
        if val_str not in field_order:
            field_order.append(val_str)
            field_rows.append((r, val_str))

    lectures = {}
    lecture_count = 1
    for col in range(2, max_col + 1):
        header_val = ws.cell(1, col).value
        col_name = str(header_val).strip() if header_val is not None else ''

        values = {}
        for r, key in field_rows:
            cell = ws.cell(r, col)
            val = cell.value
            if isinstance(val, (datetime, date)):
                values[key] = val
            elif val is None:
                values[key] = ''
            else:
                values[key] = str(val)

        if any(str(v).strip() for v in values.values()):
            name = ''
            if col_name and col_name.upper() not in ('NỘI DUNG', 'TÊN TRƯỜNG THÔNG TIN', 'NHÃN BÀI GIẢNG'):
                name = col_name
            elif values.get('Tên bài giảng'):
                tb = str(values.get('Tên bài giảng')).strip()
                if tb:
                    name = tb.split('\n')[0].strip()

            if not name:
                name = f'Bài {lecture_count}'

            base_name = name
            dup_idx = 1
            while name in lectures:
                dup_idx += 1
                name = f"{base_name} ({dup_idx})"
            lectures[name] = values
            lecture_count += 1

    return field_order, lectures


# ---------------------------------------------------------------- Định dạng ngày --
def format_date(value, fmt):
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return ''
        try:
            value = datetime.strptime(value, '%d/%m/%Y')
        except ValueError:
            try:
                value = datetime.strptime(value, '%Y-%m-%d')
            except ValueError:
                return value  # không parse được, trả nguyên văn
    day = value.day
    month = value.month
    year = value.year
    month_s = f'{month:02d}' if month in (1, 2) else str(month)
    day_s = f'{day:02d}' if day < 10 else str(day)

    out = fmt.replace('yyyy', str(year)).replace('dd', day_s)
    # Chỉ thay ký hiệu "m" đứng riêng một mình (bọc khoảng trắng/đầu-cuối chuỗi),
    # để không đụng vào chữ "m" nằm sẵn trong các từ như "năm", "tháng".
    out = re.sub(r'(?<![^\s])m(?![^\s])', month_s, out)
    return out


def coerce_value(name, fmt, raw):
    kind = field_kind(name, fmt)
    if kind == 'date':
        return format_date(raw, fmt)
    if kind == 'number':
        try:
            return str(int(float(raw)))
        except (ValueError, TypeError):
            return str(raw)
    return str(raw) if raw is not None else ''


# ---------------------------------------------------------------- Gộp XML --
def _merge_one_xml(xml_bytes, values_by_placeholder):
    """
    values_by_placeholder: dict {'{{Tên trường}}': 'giá trị đã format'} hoặc
                           {'{{Tên trường:định dạng gốc}}': 'giá trị'}
    Trả về bytes XML đã thay thế, giữ nguyên toàn bộ thuộc tính run/đoạn.
    """
    root = etree.fromstring(xml_bytes)

    # Duyệt phẳng theo đúng thứ tự tài liệu: không lặp lại qua p.iter() lồng nhau
    # (tránh đếm trùng nội dung nằm trong hộp văn bản / textbox lồng bên trong đoạn khác).
    chars = []            # (char, node, local_index)
    node_para = {}
    prev_p = None
    for n in root.iter(W + 't'):
        p = n.getparent()
        while p is not None and etree.QName(p).localname != 'p':
            p = p.getparent()
        if p is not prev_p and prev_p is not None:
            chars.append(('\n', None, -1))
        prev_p = p
        node_para[id(n)] = p
        for i, c in enumerate(n.text or ''):
            chars.append((c, n, i))
    if prev_p is not None:
        chars.append(('\n', None, -1))

    full = ''.join(c for c, _, _ in chars)
    spans = list(PLACEHOLDER_RE.finditer(full))
    if not spans:
        return etree.tostring(root, xml_declaration=True, encoding='UTF-8', standalone=True)

    edits, nodes_by_id, para_touch = {}, {}, {}
    multiline_targets = []   # (node, [dòng...]) xử lý sau khi đã thay ký tự đơn giản

    for m in spans:
        key = m.group(0)
        value = values_by_placeholder.get(key)
        if value is None:
            # thử khớp theo tên trường không kèm định dạng
            value = values_by_placeholder.get('{{%s}}' % m.group(1).strip(), '')
        lines = value.split('\n') if isinstance(value, str) else [str(value)]
        repl = lines[0]

        per_node, order = {}, []
        for idx in range(m.start(), m.end()):
            c, n, li = chars[idx]
            if n is None:
                continue
            if id(n) not in per_node:
                per_node[id(n)] = [n, li, li]
                order.append(id(n))
            per_node[id(n)][2] = li
        first = True
        for nid in order:
            n, a, b = per_node[nid]
            nodes_by_id[nid] = n
            cur = edits.get(nid, list(n.text or ''))
            for k in range(a, b + 1):
                cur[k] = ''
            if first:
                cur[a] = repl
                first = False
            edits[nid] = cur
            p = node_para[nid]
            para_touch.setdefault(id(p), [p, False])
        if order:
            para_touch[id(node_para[order[0]])][1] = True
            if len(lines) > 1:
                multiline_targets.append((nodes_by_id[order[0]], lines[1:]))

    for nid, cur in edits.items():
        n = nodes_by_id[nid]
        n.text = ''.join(cur)
        n.set(XMLNS + 'space', 'preserve')

    # nhân bản đoạn cho các dòng phụ (Alt+Enter trong Excel)
    for t_node, extra_lines in multiline_targets:
        p = t_node.getparent()
        while p is not None and etree.QName(p).localname != 'p':
            p = p.getparent()
        anchor = p
        for line in extra_lines:
            p2 = copy.deepcopy(p)
            for t2 in p2.iter(W + 't'):
                t2.text = line
                t2.set(XMLNS + 'space', 'preserve')
                break
            anchor.addnext(p2)
            anchor = p2
            para_touch[id(p2)] = [p2, True]  # đoạn mới không được xóa

    removed = 0
    for p, is_first in list(para_touch.values()):
        if is_first:
            continue
        # Không xóa nếu đoạn văn chứa ngắt phần (sectPr) để tránh hỏng viền trang (page border)
        if p.find('.//' + W + 'sectPr') is not None:
            continue
        # Không xóa nếu đoạn chứa ngắt trang (br type="page")
        has_page_break = False
        for br in p.iter(W + 'br'):
            if br.get(W + 'type') == 'page':
                has_page_break = True
                break
        if has_page_break:
            continue
            
        if ''.join(n.text or '' for n in p.iter(W + 't')).strip() == '' and p.getparent() is not None:
            p.getparent().remove(p)
            removed += 1

    return etree.tostring(root, xml_declaration=True, encoding='UTF-8', standalone=True)


def merge_docx(template_path, field_values_raw, fields_meta, field_formats, output_path):
    """
    field_values_raw: dict {ten_truong: gia_tri_tho (str hoặc datetime)}
    fields_meta: list [(ten_truong, dinh_dang_dai_dien)] lấy từ extract_fields()
    field_formats: {ten_truong: {dinh_dang1, dinh_dang2, ...}} lấy từ extract_fields() —
        mỗi định dạng ứng với một vị trí xuất hiện riêng của trường trong văn bản, để
        cùng một dữ liệu nhập vào có thể hiển thị đúng theo từng kiểu khác nhau ở từng chỗ
        (ví dụ "Ngày biên soạn" vừa cần "THÁNG m NĂM yyyy" vừa cần đủ "Ngày dd tháng m năm yyyy").
    """
    fmt_of = dict(fields_meta)
    values_by_placeholder = {}
    for name, raw in field_values_raw.items():
        default_fmt = fmt_of.get(name)
        values_by_placeholder['{{%s}}' % name] = coerce_value(name, default_fmt, raw)
        for fmt in field_formats.get(name, ()):
            values_by_placeholder['{{%s:%s}}' % (name, fmt)] = coerce_value(name, fmt, raw)

    tmp_dir = tempfile.mkdtemp()
    try:
        with zipfile.ZipFile(template_path) as zin:
            names = zin.namelist()
            with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zout:
                for n in names:
                    data = zin.read(n)
                    if n == 'word/document.xml' or re.match(r'word/(header|footer)\d*\.xml', n):
                        data = _merge_one_xml(data, values_by_placeholder)
                    zout.writestr(n, data)
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)
    return output_path


from openpyxl import Workbook
from openpyxl.styles import Protection, PatternFill, Font

def generate_protected_excel(docx_path, out_xlsx_path, num_lectures=10):
    fields_meta, _ = extract_fields(docx_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "DuLieu_BaiGiang"
    
    # Enable protection
    ws.protection.sheet = True
    ws.protection.enable()

    # Style for headers
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True)
    field_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    field_font = Font(bold=True)

    # Column A: Field names
    ws.cell(row=1, column=1, value="Tên trường (Không sửa)").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).protection = Protection(locked=True)
    ws.column_dimensions['A'].width = 35

    for row_idx, (field_name, _) in enumerate(fields_meta, start=2):
        cell = ws.cell(row=row_idx, column=1, value=field_name)
        cell.font = field_font
        cell.fill = field_fill
        cell.protection = Protection(locked=True)

    # Columns B to Z: Lectures
    for col_idx in range(2, num_lectures + 2):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = 40
        
        # Header (Lecture name)
        header_cell = ws.cell(row=1, column=col_idx, value=f"Bài giảng {col_idx - 1}")
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.protection = Protection(locked=False) # Unlock for editing lecture name

        # Data cells
        for row_idx in range(2, len(fields_meta) + 2):
            data_cell = ws.cell(row=row_idx, column=col_idx)
            data_cell.protection = Protection(locked=False) # Unlock for data entry
            
    wb.save(out_xlsx_path)

# ---------------------------------------------------------------- Xuất PDF --
def _find_soffice():
    import sys
    base = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
    paths = [
        os.path.join(base, 'LibreOffice', 'program', 'soffice.exe'),
        os.path.join(os.getcwd(), 'LibreOffice', 'program', 'soffice.exe'),
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe"
    ]
    for p in paths:
        if os.path.exists(p):
            return p
    return 'soffice'

def docx_to_pdf(docx_path, output_dir):
    soffice = _find_soffice()
    flags = subprocess.CREATE_NO_WINDOW if os.name == 'nt' else 0
    pdf_name = os.path.splitext(os.path.basename(docx_path))[0] + '.pdf'
    out_pdf_path = os.path.join(output_dir, pdf_name)
    
    if os.path.exists(out_pdf_path):
        try:
            os.remove(out_pdf_path)
        except Exception:
            pass

    cmd = [soffice, '--headless', '--convert-to', 'pdf', docx_path, '--outdir', output_dir]
    subprocess.run(cmd, check=True, creationflags=flags)
    return out_pdf_path
