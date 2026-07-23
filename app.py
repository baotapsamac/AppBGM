# -*- coding: utf-8 -*-
import os
import io
import sys
import json
import uuid
import atexit
import shutil
import zipfile
import tempfile
import threading
import webbrowser
import webview
from flask import Flask, request, jsonify, send_file, render_template

import merge


def resource_path(relative):
    """Trả về đường dẫn đúng cả khi chạy bằng 'python app.py' lẫn khi đã đóng gói .exe
    (PyInstaller giải nén tài nguyên vào một thư mục tạm gọi là sys._MEIPASS)."""
    base = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, relative)


app = Flask(
    __name__,
    template_folder=resource_path('templates'),
    static_folder=resource_path('static'),
)
BASE = tempfile.mkdtemp(prefix='baigiang_')
atexit.register(shutil.rmtree, BASE, ignore_errors=True)

# Mẫu Word cố định, đóng gói sẵn cùng công cụ — người dùng chỉ cần nạp file
# Excel dữ liệu, không cần (và không thể) đổi sang mẫu Word khác.
TEMPLATE_PATH = resource_path('sample/MAU_BaiGiang.docx')
TEMPLATE_FIELDS_META, TEMPLATE_FIELD_FORMATS = merge.extract_fields(TEMPLATE_PATH)


def _fields_for_ui(field_order):
    fmt_of = dict(TEMPLATE_FIELDS_META)
    order = field_order or [f for f, _ in TEMPLATE_FIELDS_META]
    out = []
    for name in order:
        fmt = fmt_of.get(name)
        out.append({'name': name, 'kind': merge.field_kind(name, fmt), 'format': fmt})
    return out


@app.route('/')
def index():
    return render_template('index.html')


def _parse_excel(xlsx_path):
    try:
        field_order, lectures = merge.read_lectures(xlsx_path)
    except Exception as e:
        return {'error': 'Lỗi khi đọc file: %s' % e}, 400

    if not lectures:
        return {'error': 'Không tìm thấy cột bài giảng nào trong file Excel.'}, 400

    # Merge legacy fields if present (Tài liệu học tập chính, Tài liệu tham khảo)
    # into 'Vật chất bảo đảm của học viên'
    for name, vals in lectures.items():
        tl1 = vals.pop('Tài liệu học tập chính', None)
        tl2 = vals.pop('Tài liệu tham khảo', None)
        if tl1 or tl2:
            base = vals.get('Vật chất bảo đảm của học viên', '')
            merged = []
            if tl1: merged.append('- Tài liệu học tập chính: ' + str(tl1))
            if tl2: merged.append('- Tài liệu học tập tham khảo: ' + str(tl2))
            if base: merged.append(str(base))
            vals['Vật chất bảo đảm của học viên'] = '\n'.join(merged)
            
    field_order = [f for f in field_order if f not in ('Tài liệu học tập chính', 'Tài liệu tham khảo')]

    tpl_names = {f for f, _ in TEMPLATE_FIELDS_META}
    xlsx_names = set(field_order)
    missing_in_xlsx = sorted(tpl_names - xlsx_names)
    extra_in_xlsx = sorted(xlsx_names - tpl_names)

    return {
        'fields': _fields_for_ui(field_order),
        'lectures': {
            name: {k: (v.strftime('%d/%m/%Y') if hasattr(v, 'strftime') else v)
                   for k, v in vals.items()}
            for name, vals in lectures.items()
        },
        'lecture_order': list(lectures.keys()),
        'warnings': {
            'thieu_trong_excel': missing_in_xlsx,
            'thua_trong_excel': extra_in_xlsx,
        },
    }, 200


@app.route('/api/upload', methods=['POST'])
def upload():
    xlsx_file = request.files.get('xlsx')
    if not xlsx_file:
        return jsonify({'error': 'Cần tải lên file Excel dữ liệu.'}), 400

    xlsx_path = os.path.join(BASE, 'data_%s.xlsx' % uuid.uuid4().hex[:8])
    xlsx_file.save(xlsx_path)

    try:
        data, status = _parse_excel(xlsx_path)
        return jsonify(data), status
    finally:
        if os.path.exists(xlsx_path):
            os.remove(xlsx_path)


@app.route('/api/open_excel_native', methods=['POST'])
def open_excel_native():
    try:
        import webview
        if not webview.windows:
            return jsonify({'error': 'Không tìm thấy cửa sổ ứng dụng để mở hộp thoại.'}), 400
            
        file_types = ('Excel Files (*.xlsx;*.xls)', 'All files (*.*)')
        result = webview.windows[0].create_file_dialog(webview.OPEN_DIALOG, directory='', file_types=file_types)
        
        if not result or not result[0]:
            return jsonify({'success': False, 'message': 'Đã hủy chọn file.'})
            
        file_path = result[0]
        data, status = _parse_excel(file_path)
        
        if status == 200:
            data['success'] = True
            
        return jsonify(data), status
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/download_template', methods=['GET'])
def download_template():
    out_xlsx = os.path.join(BASE, 'Mau_DuLieu_%s.xlsx' % uuid.uuid4().hex[:8])
    try:
        sample_path = resource_path('sample/DULIEU_BaiGiang.xlsx')
        import openpyxl
        from openpyxl.styles import Protection
        wb = openpyxl.load_workbook(sample_path)
        ws = wb.active
        ws.protection.sheet = True
        ws.protection.enable()
        ws.protection.insertColumns = False
        ws.protection.insertRows = True
        ws.protection.deleteRows = True
        ws.protection.deleteColumns = True
        ws.protection.formatCells = False
        
        # Unlock columns B to AZ for rows 1 to 51
        for col_idx in range(2, 52):
            for row_idx in range(1, 52):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.protection = Protection(locked=False)
                
        wb.save(out_xlsx)
        with open(out_xlsx, 'rb') as f:
            data = f.read()
        return send_file(io.BytesIO(data), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='Mau_DuLieu.xlsx')
    finally:
        if os.path.exists(out_xlsx):
            os.remove(out_xlsx)

@app.route('/api/download_template_native', methods=['POST'])
def download_template_native():
    try:
        import webview
        if not webview.windows:
            return jsonify({'error': 'Không tìm thấy cửa sổ ứng dụng để mở hộp thoại.'}), 400
            
        file_types = ('Excel Files (*.xlsx)', 'All files (*.*)')
        result = webview.windows[0].create_file_dialog(webview.SAVE_DIALOG, directory='', save_filename='Mau_DuLieu.xlsx', file_types=file_types)
        
        if not result or not result[0]:
            return jsonify({'success': False, 'message': 'Đã hủy lưu file.'})
            
        save_path = result[0]
        
        sample_path = resource_path('sample/DULIEU_BaiGiang.xlsx')
        import openpyxl
        from openpyxl.styles import Protection
        wb = openpyxl.load_workbook(sample_path)
        ws = wb.active
        ws.protection.sheet = True
        ws.protection.enable()
        ws.protection.insertColumns = False
        ws.protection.insertRows = True
        ws.protection.deleteRows = True
        ws.protection.deleteColumns = True
        ws.protection.formatCells = False
        
        for col_idx in range(2, 52):
            for row_idx in range(1, 52):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.protection = Protection(locked=False)
                
        wb.save(save_path)
        return jsonify({'success': True, 'message': 'Đã lưu file mẫu thành công tại:\n' + save_path})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def _build_docx(lecture_values, out_path):
    merge.merge_docx(TEMPLATE_PATH, lecture_values, TEMPLATE_FIELDS_META, TEMPLATE_FIELD_FORMATS, out_path)


@app.route('/api/preview', methods=['POST'])
def preview():
    if request.is_json:
        payload = request.get_json()
    else:
        payload_str = request.form.get('payload')
        payload = json.loads(payload_str) if payload_str else {}
        
    values = payload.get('values', {})
    out_docx = os.path.join(BASE, 'preview_%s.docx' % uuid.uuid4().hex[:8])
    pdf_path = None
    try:
        _build_docx(values, out_docx)
        pdf_path = merge.docx_to_pdf(out_docx, BASE)
        with open(pdf_path, 'rb') as f:
            pdf_bytes = f.read()
    except Exception as e:
        return jsonify({'error': 'Lỗi khi tạo bản xem trước: %s' % e}), 500
    finally:
        # Đọc hẳn nội dung vào bộ nhớ trước rồi mới xóa file tạm ở đây, vì trên
        # Windows không thể xóa file khi nó vẫn đang được giữ mở để gửi qua send_file.
        for p in (out_docx, pdf_path):
            if p and os.path.exists(p):
                os.remove(p)

    return send_file(io.BytesIO(pdf_bytes), mimetype='application/pdf')


def _safe_name(raw_name, fallback):
    name = (raw_name or fallback).strip() or fallback
    safe = ''.join(c for c in name if c.isalnum() or c in ' _-').strip().replace(' ', '_')[:80]
    return safe or fallback


def _build_output(values, tmp_prefix):
    """Dựng 1 bài giảng, trả về (bytes, extension).
    Luôn dọn sạch mọi file tạm trước khi trả về, kể cả khi có lỗi.
    """
    out_docx = os.path.join(BASE, '%s.docx' % tmp_prefix)
    try:
        _build_docx(values, out_docx)
        with open(out_docx, 'rb') as f:
            return f.read(), 'docx'
    finally:
        if out_docx and os.path.exists(out_docx):
            os.remove(out_docx)


@app.route('/api/export', methods=['POST'])
def export_one():
    payload = request.get_json()
    values = payload.get('values', {})
    safe = _safe_name(values.get('Tên bài giảng'), 'BaiGiang')
    try:
        data, ext = _build_output(values, uuid.uuid4().hex[:8])
    except Exception as e:
        return jsonify({'error': 'Lỗi khi xuất file: %s' % e}), 500

    return send_file(io.BytesIO(data), as_attachment=True, mimetype=None,
                      download_name='%s.%s' % (safe, ext))

@app.route('/api/export_native', methods=['POST'])
def export_one_native():
    payload = request.get_json()
    values = payload.get('values', {})
    safe = _safe_name(values.get('Tên bài giảng'), 'BaiGiang')
    
    try:
        import webview
        if not webview.windows:
            return jsonify({'error': 'Không tìm thấy cửa sổ ứng dụng để mở hộp thoại.'}), 400
            
        file_types = ('Word Documents (*.docx)', 'All files (*.*)')
        result = webview.windows[0].create_file_dialog(webview.SAVE_DIALOG, directory='', save_filename='%s.docx' % safe, file_types=file_types)
        
        if not result or not result[0]:
            return jsonify({'success': False, 'message': 'Đã hủy lưu file.'})
            
        save_path = result[0]
        data, ext = _build_output(values, uuid.uuid4().hex[:8])
        with open(save_path, 'wb') as f:
            f.write(data)
        
        return jsonify({'success': True, 'message': 'Đã xuất và lưu bài giảng thành công tại:\\n' + save_path})
    except Exception as e:
        return jsonify({'error': 'Lỗi khi xuất file: %s' % e}), 500


# ---- Xuất tất cả: chạy nền + báo tiến trình, 1 bài lỗi không làm hỏng cả gói ----
EXPORT_JOBS = {}
EXPORT_JOBS_LOCK = threading.Lock()


@app.route('/api/export_all/start', methods=['POST'])
def export_all_start():
    payload = request.get_json()
    all_values = payload.get('lectures', {})  # {ten_bai: {truong: giatri}}
    if not all_values:
        return jsonify({'error': 'Chưa chọn bài giảng nào để xuất.'}), 400

    job_id = uuid.uuid4().hex[:12]
    with EXPORT_JOBS_LOCK:
        EXPORT_JOBS[job_id] = {
            'done': 0, 'total': len(all_values), 'status': 'running',
            'zip_path': None, 'error': None, 'skipped': [],
        }

    args = (job_id, all_values)
    threading.Thread(target=_run_export_all_job, args=args, daemon=True).start()
    return jsonify({'job_id': job_id})


def _run_export_all_job(job_id, all_values):
    zip_path = os.path.join(BASE, 'BoBaiGiang_%s.zip' % uuid.uuid4().hex[:8])
    used_names = set()
    skipped = []
    try:
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for lecture_name, values in all_values.items():
                safe = _safe_name(values.get('Tên bài giảng') or lecture_name, lecture_name or 'BaiGiang')
                out_docx = os.path.join(BASE, '%s.docx' % uuid.uuid4().hex[:8])
                pdf_path = None
                try:
                    merge.merge_docx(TEMPLATE_PATH, values, TEMPLATE_FIELDS_META, TEMPLATE_FIELD_FORMATS, out_docx)
                    arcname = safe
                    n = 2
                    while arcname in used_names:
                        arcname = '%s_%d' % (safe, n)
                        n += 1
                    zf.write(out_docx, arcname='%s.docx' % arcname)
                    used_names.add(arcname)
                except Exception as e:
                    skipped.append({'lecture': lecture_name, 'error': str(e)})
                finally:
                    if out_docx and os.path.exists(out_docx):
                        os.remove(out_docx)
                with EXPORT_JOBS_LOCK:
                    EXPORT_JOBS[job_id]['done'] += 1

        with EXPORT_JOBS_LOCK:
            if skipped and len(skipped) == len(all_values):
                EXPORT_JOBS[job_id]['status'] = 'error'
                EXPORT_JOBS[job_id]['error'] = 'Không xuất được bài giảng nào (%d/%d bài lỗi).' % (
                    len(skipped), len(all_values))
                EXPORT_JOBS[job_id]['skipped'] = skipped
            else:
                EXPORT_JOBS[job_id]['status'] = 'done'
                EXPORT_JOBS[job_id]['zip_path'] = zip_path
                EXPORT_JOBS[job_id]['skipped'] = skipped
        if skipped and len(skipped) == len(all_values) and os.path.exists(zip_path):
            os.remove(zip_path)
    except Exception as e:
        if os.path.exists(zip_path):
            os.remove(zip_path)
        with EXPORT_JOBS_LOCK:
            EXPORT_JOBS[job_id]['status'] = 'error'
            EXPORT_JOBS[job_id]['error'] = str(e)


@app.route('/api/export_all/progress/<job_id>')
def export_all_progress(job_id):
    with EXPORT_JOBS_LOCK:
        job = EXPORT_JOBS.get(job_id)
    if not job:
        return jsonify({'error': 'Không tìm thấy tiến trình xuất.'}), 404
    return jsonify({k: job[k] for k in ('done', 'total', 'status', 'error', 'skipped')})


@app.route('/api/export_all/result/<job_id>')
def export_all_result(job_id):
    with EXPORT_JOBS_LOCK:
        job = EXPORT_JOBS.pop(job_id, None)
    if not job or job['status'] != 'done':
        return jsonify({'error': 'Chưa có kết quả xuất (hoặc đã lấy kết quả này rồi).'}), 400

    zip_path = job['zip_path']
    try:
        with open(zip_path, 'rb') as f:
            zip_bytes = f.read()
    finally:
        if os.path.exists(zip_path):
            os.remove(zip_path)

    return send_file(io.BytesIO(zip_bytes), mimetype='application/zip', as_attachment=True, download_name='BoBaiGiang.zip')

@app.route('/api/export_all/result_native/<job_id>', methods=['POST'])
def export_all_result_native(job_id):
    with EXPORT_JOBS_LOCK:
        job = EXPORT_JOBS.pop(job_id, None)
    if not job or job['status'] != 'done':
        return jsonify({'error': 'Chưa có kết quả xuất (hoặc đã lấy kết quả này rồi).'}), 400

    zip_path = job['zip_path']
    try:
        import webview
        import shutil
        if not webview.windows:
            return jsonify({'error': 'Không tìm thấy cửa sổ ứng dụng để mở hộp thoại.'}), 400
            
        file_types = ('ZIP Files (*.zip)', 'All files (*.*)')
        result = webview.windows[0].create_file_dialog(webview.SAVE_DIALOG, directory='', save_filename='BoBaiGiang.zip', file_types=file_types)
        
        if not result or not result[0]:
            return jsonify({'success': False, 'message': 'Đã hủy lưu file.'})
            
        save_path = result[0]
        shutil.copy2(zip_path, save_path)
        return jsonify({'success': True, 'message': 'Đã xuất và lưu file thành công tại:\n' + save_path})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if os.path.exists(zip_path):
            os.remove(zip_path)


ACTIVE_EDIT_FILES = {}

def get_swriter_path():
    base_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(base_dir, 'LibreOffice', 'program', 'swriter.exe'),
        os.path.join(os.getcwd(), 'LibreOffice', 'program', 'swriter.exe'),
        resource_path('LibreOffice/program/swriter.exe'),
    ]
    for c in candidates:
        if os.path.exists(c):
            return c
    return 'swriter'

@app.route('/api/direct_edit/open', methods=['POST'])
def direct_edit_open():
    payload = request.get_json() or {}
    values = payload.get('values', {})
    
    edit_id = uuid.uuid4().hex[:8]
    out_docx = os.path.join(BASE, 'edit_%s.docx' % edit_id)
    try:
        _build_docx(values, out_docx)
    except Exception as e:
        return jsonify({'error': 'Lỗi khi tạo file bài giảng: %s' % str(e)}), 500
    
    swriter_path = get_swriter_path()
        
    try:
        import subprocess
        subprocess.Popen([swriter_path, out_docx])
        with EXPORT_JOBS_LOCK:
            ACTIVE_EDIT_FILES[edit_id] = out_docx
        return jsonify({'success': True, 'edit_id': edit_id})
    except Exception as e:
        return jsonify({'error': 'Không thể mở LibreOffice Writer (%s): %s' % (swriter_path, str(e))}), 500


def _send_save_and_close_to_libreoffice():
    try:
        import ctypes
        import time
        user32 = ctypes.windll.user32
        hwnds = []
        def enum_cb(h, extra):
            if user32.IsWindowVisible(h):
                length = user32.GetWindowTextLengthW(h)
                if length > 0:
                    buff = ctypes.create_unicode_buffer(length + 1)
                    user32.GetWindowTextW(h, buff, length + 1)
                    if 'LibreOffice Writer' in buff.value or 'edit_' in buff.value:
                        hwnds.append(h)
            return True

        WNDENUMPROC = ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.c_int, ctypes.c_int)
        user32.EnumWindows(WNDENUMPROC(enum_cb), 0)

        for h in hwnds:
            try:
                user32.SetForegroundWindow(h)
                time.sleep(0.1)
                VK_CONTROL = 0x11
                VK_S = 0x53
                KEYEVENTF_KEYUP = 0x0002
                user32.keybd_event(VK_CONTROL, 0, 0, 0)
                user32.keybd_event(VK_S, 0, 0, 0)
                time.sleep(0.05)
                user32.keybd_event(VK_S, 0, KEYEVENTF_KEYUP, 0)
                user32.keybd_event(VK_CONTROL, 0, KEYEVENTF_KEYUP, 0)
                time.sleep(0.3)
                WM_CLOSE = 0x0010
                user32.PostMessageW(h, WM_CLOSE, 0, 0)
            except Exception:
                pass
    except Exception:
        pass


@app.route('/api/direct_edit/complete', methods=['POST'])
def direct_edit_complete():
    payload = request.get_json() or {}
    edit_id = payload.get('edit_id')
    out_docx = None
    with EXPORT_JOBS_LOCK:
        out_docx = ACTIVE_EDIT_FILES.get(edit_id)
        
    if not out_docx or not os.path.exists(out_docx):
        return jsonify({'error': 'Không tìm thấy file đang chỉnh sửa.'}), 400

    import time
    _send_save_and_close_to_libreoffice()
    time.sleep(0.5)
        
    pdf_path = None
    try:
        pdf_path = merge.docx_to_pdf(out_docx, BASE)
        with open(pdf_path, 'rb') as f:
            pdf_bytes = f.read()
        return send_file(io.BytesIO(pdf_bytes), mimetype='application/pdf')
    except Exception as e:
        return jsonify({'error': 'Lỗi khi cập nhật bản xem trước: %s' % str(e)}), 500
    finally:
        if pdf_path and os.path.exists(pdf_path):
            os.remove(pdf_path)


def run_flask():
    app.run(host='127.0.0.1', port=5050, debug=False, threaded=True)

if __name__ == '__main__':
    for stream in (sys.stdout, sys.stderr):
        if hasattr(stream, 'reconfigure'):
            stream.reconfigure(encoding='utf-8', errors='replace')

    URL = 'http://127.0.0.1:5050'
    print('=' * 56)
    print(' CÔNG CỤ SOẠN BÀI GIẢNG')
    print(' Đang khởi động cửa sổ ứng dụng...')
    print('=' * 56)
    
    t = threading.Thread(target=run_flask)
    t.daemon = True
    t.start()
    
    webview.create_window('AppBGM', URL, width=1366, height=768)
    webview.start(gui='edgechromium')
